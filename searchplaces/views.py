import requests, os, time
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from django.shortcuts import render
from django.http import HttpResponse
from .forms import PlaceSearchForm

def get_list_of_detailed_places_per_search_type(location, radius, search_type, api_key):
    url = f'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={location}&radius={radius}&type={search_type}&key={api_key}'
    place_details = []
    
    while url:
        response = requests.get(url)
        data = response.json()
        places = data.get('results', [])
        
        # Extract details for each place
        for place in places:
            place_id = place.get('place_id')
            name = place.get('name', 'N/A')
            vicinity = place.get('vicinity', 'N/A')
            maps_url = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
            rating = place.get('rating', 'N/A')
            user_ratings_total = place.get('user_ratings_total', 'N/A')
            business_status = place.get('business_status', 'N/A')
            price_level = place.get('price_level', 'N/A')
            
            # Use Place Details API for additional fields
            details_url = f'https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&key={api_key}'
            details_response = requests.get(details_url)
            details_data = details_response.json().get('result', {})

            formatted_address = details_data.get('formatted_address', 'N/A')
            phone_number = details_data.get('international_phone_number') or details_data.get('formatted_phone_number', 'N/A')
            website = details_data.get('website', 'N/A')
            opening_hours = details_data.get('opening_hours', {}).get('weekday_text', [])
            open_now = details_data.get('opening_hours', {}).get('open_now', 'N/A')
            
            # Format opening hours as a single text
            opening_hours_text = "\n".join(opening_hours)

            # Append each place's details as a dictionary
            place_details.append({
                "Place Name": name,
                "Vicinity": vicinity,
                "Maps URL": maps_url,
                "Formatted Address": formatted_address,
                "Phone Number": phone_number,
                "Website": website,
                "Opening Hours": opening_hours_text,
                "Open Now": open_now,
                "Rating": rating,
                "User Ratings Total": user_ratings_total,
                "Business Status": business_status,
                "Price Level": price_level,
            })
            
        # Check for next_page_token
        next_page_token = data.get('next_page_token')
        if next_page_token:
            url = f'https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken={next_page_token}&key={api_key}'
            time.sleep(2)  # API requires delay before using next page token
        else:
            url = None
            
    return place_details

def save_to_excel_in_memory(place_data, search_type, headers, workbook, output):
    # Create an in-memory Excel workbook
    # workbook = Workbook()
    sheet = workbook.create_sheet(title=search_type)

    # Insert headers based on user selection
    sheet.append(headers)
    
    # Style the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin"), right=Side(style="thin"))

    # Insert place data based on selected headers
    for place in place_data:
        row_data = [place.get(header, 'N/A') for header in headers]
        sheet.append(row_data)

    # Set fixed column widths for readability
    column_widths = {header: 35 for header in headers}
    for col_num, (header, width) in enumerate(column_widths.items(), start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

    # Set minimum row height and wrap text
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        sheet.row_dimensions[row[0].row].height = 55

    # Save the workbook to the in-memory output
    workbook.save(output)
    output.seek(0)  # Rewind the in-memory output for reading
    return output


def place_search_view(request):
    if request.method == 'POST':
        form = PlaceSearchForm(request.POST)
        if form.is_valid():
            selected_search_types = form.cleaned_data['search_types']
            address = f"{form.cleaned_data['latitude']},{form.cleaned_data['longitude']}"
            radius = form.cleaned_data['radius']
            selected_headers = form.cleaned_data['headers']
            
            api_key = os.getenv('GOOGLE_MAPS_API_KEY')
            
            # Create an in-memory output for the entire workbook
            output = BytesIO()
            workbook = Workbook()
            
            for search_type in selected_search_types:
                # Get the place data for each search type
                place_data = get_list_of_detailed_places_per_search_type(address, radius, search_type, api_key)
                search_type_output = save_to_excel_in_memory(place_data, search_type, selected_headers, workbook, output)
            
            # Remove the default sheet if unused
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
                
            # Finalize and download the Excel file
            response = HttpResponse(search_type_output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=places_data.xlsx'
            return response
        
    else:
        form = PlaceSearchForm()

    return render(request, 'searchplaces/home_landing.html', {'form': form, 'google_maps_api_key': os.getenv('GOOGLE_MAPS_API_KEY')})
