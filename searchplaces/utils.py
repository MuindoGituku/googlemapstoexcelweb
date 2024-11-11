import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
import time
import os
from io import BytesIO

# API Key stored in environment variables
API_KEY = os.getenv('GOOGLE_MAPS_API_KEY')

def fetch_coordinates(address: str) -> str:
    """
    Fetches latitude and longitude for a given address using the Geocoding API.
    """
    geocode_url = f'https://maps.googleapis.com/maps/api/geocode/json?address={address}&key={API_KEY}'
    response = requests.get(geocode_url)
    data = response.json()
    if data['results']:
        location = data['results'][0]['geometry']['location']
        return f"{location['lat']},{location['lng']}"
    else:
        raise ValueError("Address not found")

# At the end of fetch_places_with_pagination
def fetch_places_with_pagination(address, radius, search_types, selected_headers):
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet
    
    location = fetch_coordinates(address)
    
    for search_type in search_types:
        url = f'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location={location}&radius={radius}&type={search_type}&key={API_KEY}'
        place_details = []
        
        # Fetch and process places data for each search type
        while url:
            response = requests.get(url)
            if response.status_code != 200:
                print(f"Error fetching data for search type '{search_type}': {response.text}")
                break
            data = response.json()
            places = data.get('results', [])

            for place in places:
                place_id = place.get('place_id')
                place_info = {}

                # Fetch details based on selected headers
                if 'Place Name' in selected_headers:
                    place_info['Place Name'] = place.get('name', 'N/A')
                if 'Vicinity' in selected_headers:
                    place_info['Vicinity'] = place.get('vicinity', 'N/A')
                if 'Maps URL' in selected_headers:
                    place_info['Maps URL'] = f"https://www.google.com/maps/place/?q=place_id:{place_id}"
                if 'Rating' in selected_headers:
                    place_info['Rating'] = place.get('rating', 'N/A')
                if 'User Ratings Total' in selected_headers:
                    place_info['User Ratings Total'] = place.get('user_ratings_total', 'N/A')
                if 'Types' in selected_headers:
                    place_info['Types'] = ", ".join(place.get('types', []))
                if 'Business Status' in selected_headers:
                    place_info['Business Status'] = place.get('business_status', 'N/A')
                if 'Price Level' in selected_headers:
                    place_info['Price Level'] = place.get('price_level', 'N/A')

                # Use Place Details API for additional data
                details_url = f'https://maps.googleapis.com/maps/api/place/details/json?place_id={place_id}&key={API_KEY}'
                details_response = requests.get(details_url)
                details_data = details_response.json().get('result', {})

                if 'Formatted Address' in selected_headers:
                    place_info['Formatted Address'] = details_data.get('formatted_address', 'N/A')
                if 'Phone Number' in selected_headers:
                    place_info['Phone Number'] = details_data.get('international_phone_number') or details_data.get('formatted_phone_number', 'N/A')
                if 'Website' in selected_headers:
                    place_info['Website'] = details_data.get('website', 'N/A')
                if 'Opening Hours' in selected_headers:
                    place_info['Opening Hours'] = "\n".join(details_data.get('opening_hours', {}).get('weekday_text', []))
                if 'Open Now' in selected_headers:
                    place_info['Open Now'] = details_data.get('opening_hours', {}).get('open_now', 'N/A')

                # Append place info to the list for this search type
                place_details.append(place_info)

            # Handle pagination if there are more results
            next_page_token = data.get('next_page_token')
            if next_page_token:
                url = f'https://maps.googleapis.com/maps/api/place/nearbysearch/json?pagetoken={next_page_token}&key={API_KEY}'
                time.sleep(2)  # Google API requires a slight delay for the next token
            else:
                url = None

        # Add a new sheet for this search type
        sheet = workbook.create_sheet(title=search_type)

        # Write headers
        sheet.append(selected_headers)
        
        # Write each place's data in rows
        for place in place_details:
            row = [place.get(header, 'N/A') for header in selected_headers]
            sheet.append(row)

        # Format cells for each sheet
        header_font = Font(bold=True, size=12)
        for cell in sheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column width and apply wrap text
        for col in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = max_length
            for cell in col:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save the workbook to a BytesIO stream for download
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream.getvalue()
